from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views import generic

from .forms import LoginMentee, LoginMentor, PersonalDetailForm, ExamRecordForm, InternshipDetailForm, \
    PlacementDetailForm, ChangePassword
from . import models

import bcrypt

from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime


# Create your views here.

class LoginPage(generic.TemplateView):
    template_name = "triple_M/login.html"

    def get(self, request):
        if request.method == 'GET':
            return render(request, self.template_name)


class LoginTypePage(generic.TemplateView):
    template_name = "triple_M/login.html"

    def post(self, request, login_type):
        if request.method == 'POST':
            if login_type == "MenteeLogin":
                form = LoginMentee(request.POST)
                if form.is_valid():
                    regno = form.cleaned_data['regno']
                    password = form.cleaned_data['password']
                    # validation-code-here
                    error = validate_mentee_login(regno, password)
                    if error == "success":
                        request.session['mentee-logged-in'] = True
                        request.session['regno'] = regno
                        return redirect('triple_M:mentee-dash')
                    else:
                        args = {'form': form, 'login_type': login_type, 'error': error}
                        return render(request, self.template_name, args)
                else:
                    args = {'form': form, 'login_type': login_type}
                    return render(request, self.template_name, args)

            elif login_type == "MentorLogin":
                form = LoginMentor(request.POST)
                if form.is_valid():
                    email = form.cleaned_data['email']
                    password = form.cleaned_data['password']
                    error = validate_mentor_login(email, password)
                    if error == "success":
                        request.session['mentor-logged-in'] = True
                        request.session['email'] = email
                        return redirect('triple_M:mentor-dash')
                    else:
                        args = {'form': form, 'login_type': login_type, 'error': error}
                        return render(request, self.template_name, args)
                else:
                    args = {'form': form, 'login_type': login_type}
                    return render(request, self.template_name, args)
            else:
                return redirect('triple_M:login')

    def get(self, request, login_type):
        if request.method == 'GET':
            if login_type == "MenteeLogin":
                args = {"login_type": login_type, 'form': LoginMentee()}
                return render(request, self.template_name, args)
            elif login_type == "MentorLogin":
                args = {"login_type": login_type, 'form': LoginMentor()}
                return render(request, self.template_name, args)
            else:
                args = {"login_type": login_type}
                return render(request, self.template_name, args)


class MentorDash(generic.TemplateView):
    template_name = 'triple_M/mentor/mentor_dash.html'

    def get(self, request):
        if request.method == "GET":
            if request.session['mentor-logged-in']:
                mentor_details = list(models.Mentor.objects.filter(mentor_email=request.session["email"]).values())
                mentor_students = list(models.Admitted_Student.objects.filter(
                    mentor_mentee__mentor__mentor_email=request.session["email"]).values("reg_no", "first_name",
                                                                                         "last_name"))
                list_student = get_all_verification(mentor_students)
                verified_students = list_student[-2]
                pending_verifications = list_student[-1]
                not_filled = list_student[-3]
                request.session['mentor_details'] = mentor_details[0]
                request.session['mentor_students'] = mentor_students
                request.session['total_students'] = len(mentor_students)
                args = {'email': request.session['email'],
                        'mentor_students': request.session['mentor_students'],
                        'mentor_details': request.session['mentor_details'],
                        'total_students': request.session['total_students'],
                        'verified_students': verified_students,
                        'pending_verifications': pending_verifications,
                        'not_filled': not_filled
                        }
                return render(request, self.template_name, args)
            else:
                return redirect("triple_M:login")


class GetInternship(generic.TemplateView):
    template_name = 'triple_M/mentor/internship_details.html'

    def get(self, request):
        if request.session['mentor-logged-in']:
            intern_students = list(models.Internship_Detail.objects.filter(
                admitted_student__mentor_mentee__mentor__mentor_email="shilpa.mhatre@sakec.ac.in")
                                   .values('admitted_student__first_name',
                                           'admitted_student__last_name', 'company',
                                           'admitted_student__mentor_mentee__mentor__mentor_name',
                                           'duration_start', 'duration_end', 'certificate', 'intern_eval',
                                           'admitted_student__personal_detail__year',
                                           'admitted_student__personal_detail__division',
                                           'admitted_student__personal_detail__roll_no'))

            args = {'email': request.session['email'],
                    'mentor_students': request.session['mentor_students'],
                    'mentor_details': request.session['mentor_details'],
                    'total_students': request.session['total_students'],
                    'intern_students': intern_students
                    }
            return render(request, self.template_name, args)
        else:
            return redirect("triple_M:login")


class GetPlacement(generic.TemplateView):
    template_name = 'triple_M/mentor/placement_details.html'

    def get(self, request):
        if request.session['mentor-logged-in']:
            placed_students = list(models.Placement_Detail.objects.filter(
                admitted_student__mentor_mentee__mentor__mentor_email="shilpa.mhatre@sakec.ac.in")
                                   .values('admitted_student__first_name',
                                           'admitted_student__last_name',
                                           'admitted_student__mentor_mentee__mentor__mentor_name',
                                           'admitted_student__personal_detail__year',
                                           'admitted_student__personal_detail__division',
                                           'admitted_student__personal_detail__roll_no',
                                           'company', 'ctc', 'bond', 'placed_through', 'offer_letter'))

            args = {'email': request.session['email'],
                    'mentor_students': request.session['mentor_students'],
                    'mentor_details': request.session['mentor_details'],
                    'total_students': request.session['total_students'],
                    'placed_students': placed_students
                    }
            return render(request, self.template_name, args)
        else:
            return redirect("triple_M:login")


class VerifyDetails(generic.TemplateView):
    template_name = 'triple_M/mentor/verify_details.html'

    def get(self, request):
        if 'mentor-logged-in' in request.session:
            list_student = get_all_verification(request.session['mentor_students'])
            list_student.pop()
            list_student.pop()
            list_student.pop()
            args = {'email': request.session['email'],
                    'mentor_students': list_student,
                    'mentor_details': request.session['mentor_details'],
                    'total_students': request.session['total_students']}
            return render(request, self.template_name, args)
        else:
            return redirect("triple_M:login")


class VerifySpecificCategory(generic.TemplateView):
    template_name = 'triple_M/mentor/student_verification.html'

    def get(self, request, reg_no, section):
        if 'mentor-logged-in' in request.session:
            mentee = get_mentee(reg_no)
            if section == 'personal_detail':
                try:
                    mentee_personal_detail = models.Personal_Detail.objects.get(
                        admitted_student__reg_no=reg_no).__dict__
                    args = {'mentee': mentee, 'mentor_details': request.session['mentor_details'],
                            'personal_detail': mentee_personal_detail, 'section': section, 'reg_no': reg_no}
                    return render(request, self.template_name, args)
                except:
                    args = {'mentee': mentee, 'mentor_details': request.session['mentor_details'],
                            'reg_no': reg_no, 'section': section}
                    return render(request, self.template_name, args)
            elif section == 'exam_record':
                try:
                    mentee_personal_detail = models.Personal_Detail.objects.get(
                        admitted_student__reg_no=reg_no)
                    mentee_exam_record = models.Exam_Record.objects.get(
                        admitted_student__reg_no=reg_no).__dict__
                    del mentee_exam_record['_state']
                    del mentee_exam_record['year']
                    del mentee_exam_record['id']
                    del mentee_exam_record['admitted_student_id']
                    sem_count = get_sem_count(mentee_personal_detail.year)
                    args = {'mentee': mentee, 'mentor_details': request.session['mentor_details'],
                            'personal_detail': mentee_personal_detail,
                            'exam_record': mentee_exam_record,
                            'sem_count': sem_count, 'reg_no': reg_no, 'section': section
                            }
                    return render(request, self.template_name, args)

                except:
                    args = {'mentee': mentee, 'mentor_details': request.session['mentor_details'],
                            'reg_no': reg_no, 'section': section}
                    return render(request, self.template_name, args)
            elif section == 'internship_detail':
                try:
                    internship_list = get_internship_detail(mentee['id'])
                    args = {'mentee': mentee,
                            'mentor_details': request.session['mentor_details'],
                            'list': internship_list, 'reg_no': reg_no, 'section': section}
                    return render(request, self.template_name, args)
                except:
                    args = {'mentee': mentee, 'mentor_details': request.session['mentor_details'],
                            'reg_no': reg_no, 'section': section}
                    return render(request, self.template_name, args)
            elif section == 'placement_detail':
                try:
                    placement_list = get_placement_detail(mentee['id'])
                    args = {'mentee': mentee,
                            'mentor_details': request.session['mentor_details'],
                            'placed_list': placement_list, 'reg_no': reg_no, 'section': section}
                    return render(request, self.template_name, args)
                except:
                    args = {'mentee': mentee, 'mentor_details': request.session['mentor_details'],
                            'reg_no': reg_no, 'section': section}
                    return render(request, self.template_name, args)
        else:
            return redirect("triple_M:login")


class MentorStudent(generic.TemplateView):
    template_name = "triple_M/mentor/student.html"

    def get(self, request, reg_no):
        if 'mentor-logged-in' in request.session:
            mentee = get_mentee(reg_no)
            try:
                profile = models.Personal_Detail.objects.values('profile_photo').get(admitted_student__reg_no=reg_no)[
                    'profile_photo']
            except:
                profile = None
            verification = verify_all_details(reg_no)
            args = {'email': request.session['email'],
                    'mentee': mentee,
                    'mentor_details': request.session['mentor_details'],
                    'total_students': request.session['total_students'],
                    'reg_no': reg_no, 'verification': verification,
                    'profile_photo': profile}
            return render(request, self.template_name, args)
        else:
            return redirect("triple_M:login")


class ApproveOrDisapprove(generic.TemplateView):
    template_name = 'triple_M/mentor/student_verification.html'

    def get(self, request, reg_no, section, output, id):
        if 'mentor-logged-in' in request.session:
            if output == "disapprove":
                verified = "NO"
            elif output == "approve":
                verified = "YES"
            if section == 'personal_detail':
                mentee_personal_detail = models.Personal_Detail.objects.get(
                    admitted_student__reg_no=reg_no)
                mentee_personal_detail.verified = verified
                mentee_personal_detail.save()
                return redirect('triple_M:mentor-student', section=section, reg_no=reg_no)
            elif section == 'exam_record':
                mentee_exam_record = models.Exam_Record.objects.get(
                    admitted_student__reg_no=reg_no)
                mentee_exam_record.verified = verified
                mentee_exam_record.save()
                return redirect('triple_M:mentor-student', section=section, reg_no=reg_no)
            elif section == 'internship_detail':
                intern = models.Internship_Detail.objects.get(id=id)
                intern.verified = verified
                intern.save()
                return redirect('triple_M:mentor-student', section=section, reg_no=reg_no)
            elif section == 'placement_detail':
                placed = models.Placement_Detail.objects.get(id=id)
                placed.verified = verified
                placed.save()
                return redirect('triple_M:mentor-student', section=section, reg_no=reg_no)

        else:
            return redirect("triple_M:login")


class ContactMentee(generic.TemplateView):
    template_name = "triple_M/mentor/contact_mentee.html"

    def get(self, request):
        if 'mentor-logged-in' in request.session:
            args = {'email': request.session['email'],
                    'mentor_students': request.session['mentor_students'],
                    'mentor_details': request.session['mentor_details'],
                    'total_students': request.session['total_students'],
                    }
            return render(request, self.template_name, args)
        else:
            return redirect('triple_M:login')


class ContactSpecificMentee(generic.TemplateView):
    template_name = "triple_M/mentor/contact_mentee.html"

    def get(self, request, reg_no):
        if 'mentor-logged-in' in request.session:
            try:
                personal_detail = models.Personal_Detail.objects.values('admitted_student__first_name',
                                                                        'admitted_student__last_name', 'email', 'ph_no',
                                                                        'parent_ph_no',
                                                                        'address').get(
                    admitted_student__reg_no=reg_no)

                if personal_detail['email'] is None:
                    error = "Not Filled"
                else:
                    error = None
            except:
                error = "Not Filled"
                personal_detail = None
            args = {'email': request.session['email'],
                    'mentor_students': request.session['mentor_students'],
                    'mentor_details': request.session['mentor_details'],
                    'total_students': request.session['total_students'],
                    'personal_details': personal_detail,
                    'error': error
                    }
            return render(request, self.template_name, args)

        else:
            return redirect('triple_M:login')


class MenteeDash(generic.TemplateView):
    template_name = 'triple_M/mentee_dash.html'

    def get(self, request):
        if 'mentee-logged-in' in request.session:
            if request.method == "GET":
                mentee = get_mentee(request.session['regno'])
                mentee_personal_detail = get_mentee_personal_detail(mentee['id'])
                mentee_exam_record = get_mentee_exam_record(mentee['id'])
                mentor_name = get_mentor_name(mentee['id'])
                average = get_avg(mentee_exam_record)
                total_kt = get_kt(mentee_exam_record)
                args = {'mentee': mentee, 'mentor_name': mentor_name,
                        'personal_detail': mentee_personal_detail,
                        'exam_record': mentee_exam_record,
                        'average': average,
                        'kt': total_kt}
                return render(request, self.template_name, args)

        else:
            return redirect('triple_M:login')


class ContactMentor(generic.TemplateView):
    template_name = "triple_M/contact-mentor.html"

    def get(self, request):
        if 'mentee-logged-in' in request.session:
            mentee = get_mentee(request.session['regno'])
            mentor_name = get_mentor_name(mentee['id'])
            mentor_contact = models.Mentor.objects.values('mentor_name', 'mentor_email', 'mentor_phone_no').get(
                mentor_name=mentor_name)
            args = {'mentee': mentee, 'mentor_contact': mentor_contact, 'mentor_name': mentor_name}
            return render(request, self.template_name, args)
        else:
            return redirect('triple_M:login')


class EditDetails(generic.TemplateView):
    template_name = 'triple_M/edit_details.html'

    def get(self, request, section):
        if 'mentee-logged-in' in request.session:
            active = {'one': "", 'two': "", 'three': "", 'four': ""}
            mentee = get_mentee(request.session['regno'])
            if section == 'personal_details':
                student, created = models.Personal_Detail.objects.get_or_create(
                    admitted_student__reg_no=request.session['regno'])
                form = PersonalDetailForm(instance=student)
                active['one'] = "active"
                if 'alert' in request.session:
                    alert = request.session['alert']
                    del request.session['alert']
                    args = {'form': form, 'active': active, 'section': 'Personal Details', 'alert': alert}
                else:
                    args = {'form': form, 'active': active, 'section': 'Personal Details'}
                return render(request, self.template_name, args)
            elif section == 'exam_records':
                if models.Personal_Detail.objects.get(admitted_student__reg_no=request.session['regno']).year:
                    exam, created = models.Exam_Record.objects.get_or_create(
                        admitted_student__reg_no=request.session['regno'])
                    sem_count = get_sem_count(
                        models.Personal_Detail.objects.values('year').get(admitted_student_id=exam.admitted_student_id)[
                            'year']
                    )
                    if 'alert' in request.session:
                        del request.session['alert']

                    active['two'] = "active"

                    form = ExamRecordForm(instance=exam)

                    args = {'form': form, 'active': active, 'section': 'Exam Records', 'exam': exam,
                            'sem_count': sem_count}

                    return render(request, self.template_name, args)

                else:
                    request.session['alert'] = "Fill the Personal Details First!"
                    return redirect('triple_M:edit-details', section='personal_details')
            elif section == 'internship_details':
                internship_list = get_internship_detail(mentee['id'])
                # args = {'mentee': mentee, 'list': internship_list}
                # form = InternshipDetailForm()
                active['three'] = "active"
                args = {'active': active, 'section': 'Internship Details', 'mentee': mentee,
                        'list': internship_list}
                return render(request, self.template_name, args)

            elif section == 'placement_details':
                placement_list = get_placement_detail(mentee['id'])
                # form = PlacementDetailForm()
                active['four'] = "active"
                args = {'active': active, 'section': 'Placement Details', 'mentee': mentee,
                        'placement_list': placement_list}
                return render(request, self.template_name, args)
            else:
                return redirect("triple_M:mentee-dash")
        else:
            return redirect('triple_M:login')

    def post(self, request, section):
        if request.method == 'POST':
            active = {'one': "", 'two': "", 'three': "", 'four': ""}
            if section == 'personal_details':
                active['one'] = "active"
                student = models.Personal_Detail.objects.get(admitted_student__reg_no=request.session['regno'])
                form = PersonalDetailForm(request.POST, request.FILES, instance=student)
                if form.is_valid():
                    error = check_personal_detail_form(form.cleaned_data)
                    if error != 'pass':
                        args = {'form': form, 'active': active, 'section': 'Personal Details',
                                'alert': 'There was some error in your Submission',
                                'error': error}
                        return render(request, self.template_name, args)
                    else:
                        if form.has_changed():
                            print('YES')
                            student.verified = "NO"
                            student.save()
                        form.save()
                        return redirect('triple_M:edit-details', section="exam_records")

                else:
                    args = {'form': form, 'active': active, 'section': 'Personal Details',
                            'alert': 'There was some error in your Submission'}
                    return render(request, self.template_name, args)
            elif section == 'exam_records':
                active['two'] = "active"

                exam = models.Exam_Record.objects.get(admitted_student__reg_no=request.session['regno'])
                sem_count = get_sem_count(
                    models.Personal_Detail.objects.values('year').get(admitted_student_id=exam.admitted_student_id)[
                        'year']
                )

                form = ExamRecordForm(request.POST, request.FILES, instance=exam)

                if form.is_valid():
                    error = check_exam_form(form.cleaned_data, sem_count)
                    if len(error) != 0:
                        args = {'form': form, 'active': active, 'section': 'Exam Records',
                                'alert': 'There was some error in your Submission',
                                'sem_count': sem_count, 'error': error, 'exam': exam}
                        return render(request, self.template_name, args)
                    else:
                        if form.has_changed():
                            print('YES')
                            exam.verified = "NO"
                            exam.save()
                        form.save()
                        return redirect('triple_M:edit-details', section='internship_details')
                else:
                    args = {'form': form, 'active': active, 'section': 'Exam Records',
                            'alert': 'There was some error in your Submission',
                            'sem_count': sem_count, 'exam': exam}
                    return render(request, self.template_name, args)


class EditInternship(generic.TemplateView):
    template_name = "triple_M/edit_details.html"

    def get(self, request, section, id):
        if 'mentee-logged-in' in request.session:
            if section == 'internship_details':
                active = {'one': "", 'two': "", 'three': "", 'four': ""}
                mentee = get_mentee(request.session['regno'])
                active['three'] = "active"
                if id == 'new':
                    form = InternshipDetailForm()
                    args = {'form': form, 'active': active, 'section': 'Internship Details', 'mentee': mentee, 'id': id}
                    return render(request, self.template_name, args)
                else:
                    try:
                        internship = models.Internship_Detail.objects.get(id=id)
                        form = InternshipDetailForm(instance=internship)
                        args = {'form': form, 'active': active, 'section': 'Internship Details', 'mentee': mentee,
                                'id': id}
                        return render(request, self.template_name, args)
                    except:
                        return redirect('triple_M:edit-details', section="internship_details")
            elif section == 'placement_details':
                active = {'one': "", 'two': "", 'three': "", 'four': ""}
                mentee = get_mentee(request.session['regno'])
                active['four'] = "active"
                if id == 'new':
                    form = PlacementDetailForm()
                    args = {'form': form, 'active': active, 'section': 'Placement Details', 'mentee': mentee, 'id': id}
                    return render(request, self.template_name, args)
                else:
                    try:
                        placement = models.Placement_Detail.objects.get(id=id)
                        form = PlacementDetailForm(instance=placement)
                        args = {'form': form, 'active': active, 'section': 'Placement Details', 'mentee': mentee,
                                'id': id}
                        return render(request, self.template_name, args)
                    except:
                        return redirect('triple_M:edit-details', section="placement_details")
            else:
                return redirect('triple_M:edit-details', section='personal_details')
        else:
            return redirect('triple_M:login')

    def post(self, request, section, id):
        if request.method == 'POST':
            if section == 'internship_details':
                active = {'one': "", 'two': "", 'three': "", 'four': ""}
                mentee = get_mentee(request.session['regno'])
                active['three'] = "active"
                if id == "new":
                    internship = models.Internship_Detail(admitted_student_id=mentee['id'])
                else:
                    internship = models.Internship_Detail.objects.get(id=id)

                form = InternshipDetailForm(request.POST, request.FILES, instance=internship)
                if form.is_valid():
                    if form.has_changed():
                        print("YES")
                        internship.verified = "NO"
                    form.save()
                    return redirect('triple_M:edit-details', section='internship_details')
                else:
                    args = {'form': form, 'active': active, 'section': 'Internship Details',
                            'alert': "There is some error in your submission bro", 'id': id}
                    return render(request, self.template_name, args)

            elif section == 'placement_details':
                active = {'one': "", 'two': "", 'three': "", 'four': ""}
                mentee = get_mentee(request.session['regno'])
                active['four'] = "active"
                if id == "new":
                    placement = models.Placement_Detail(admitted_student_id=mentee['id'])
                else:
                    placement = models.Placement_Detail.objects.get(id=id)

                form = PlacementDetailForm(request.POST, request.FILES, instance=placement)
                if form.is_valid():
                    if form.has_changed():
                        print("YES")
                        placement.verified = "NO"
                    form.save()
                    return redirect('triple_M:edit-details', section='placement_details')
                else:
                    args = {'form': form, 'active': active, 'section': 'Placement Details',
                            'alert': "There is some error in your submission bruv", 'id': id}
                    return render(request, self.template_name, args)


class PersonalDetails(generic.TemplateView):
    template_name = "triple_M/personal_details.html"

    def get(self, request):
        if 'mentee-logged-in' in request.session:
            mentee = get_mentee(request.session['regno'])
            mentee_personal_detail = get_mentee_personal_detail(mentee['id'])
            mentor_name = get_mentor_name(mentee['id'])
            args = {'mentee': mentee, 'mentor_name': mentor_name,
                    'personal_detail': mentee_personal_detail}
            return render(request, self.template_name, args)
        else:
            return redirect('triple_M:login')


class AcademicDetails(generic.TemplateView):
    template_name = "triple_M/academic_details.html"

    def get(self, request):
        if 'mentee-logged-in' in request.session:
            mentee = get_mentee(request.session['regno'])
            mentee_personal_detail = get_mentee_personal_detail(mentee['id'])
            mentee_exam_record = get_mentee_exam_record(mentee['id'])
            mentor_name = get_mentor_name(mentee['id'])
            average = get_avg(mentee_exam_record)
            total_kt = get_kt(mentee_exam_record)
            if mentee_personal_detail['year']:
                sem_count = get_sem_count(mentee_personal_detail['year'])
            else:
                sem_count = 0
            args = {'mentee': mentee, 'mentor_name': mentor_name,
                    'personal_detail': mentee_personal_detail,
                    'exam_record': mentee_exam_record,
                    'average': average,
                    'kt': total_kt,
                    'sem_count': sem_count
                    }
            return render(request, self.template_name, args)
        else:
            return redirect('triple_M:login')


class InternshipDetails(generic.TemplateView):
    template_name = "triple_M/internship_details.html"

    def get(self, request):
        if 'mentee-logged-in' in request.session:
            mentee = get_mentee(request.session['regno'])
            intenrship_list = get_internship_detail(mentee['id'])
            verified = get_verified_details(intenrship_list)
            mentor_name = get_mentor_name(mentee['id'])
            args = {'mentee': mentee, 'mentor_name': mentor_name, 'list': intenrship_list, 'verified': verified}
            return render(request, self.template_name, args)
        else:
            return redirect('triple_M:login')


class PlacementDetails(generic.TemplateView):
    template_name = "triple_M/placement_details.html"

    def get(self, request):
        if 'mentee-logged-in' in request.session:
            mentee = get_mentee(request.session['regno'])
            placement_list = get_placement_detail(mentee['id'])
            verified = get_verified_details(placement_list)
            mentor_name = get_mentor_name(mentee['id'])
            args = {'mentee': mentee, 'mentor_name': mentor_name, 'list': placement_list, 'verified': verified}
            return render(request, self.template_name, args)
        else:
            return redirect('triple_M:login')


def change_password(request):
    if request.method == "GET":
        if 'mentee-logged-in' in request.session:
            reg_no = request.session['regno']
            form = ChangePassword()
            context = {'form': form, 'reg_no': reg_no}
            return render(request, 'triple_M/change_password.html', context)
        elif 'mentor-logged-in' in request.session:
            email = request.session['email']
            form = ChangePassword()
            context = {'form': form, 'email': email}
            return render(request, 'triple_M/change_password.html', context)
        else:
            return redirect('triple_M:login')
    if request.method == "POST":
        if 'regno' in request.session:
            reg_no = request.session['regno']
            form = ChangePassword(request.POST)
            if form.is_valid():
                mentee = models.Admitted_Student.objects.get(reg_no=reg_no)
                if check_password(form.cleaned_data['current_password'],
                                  mentee.password) or mentee.password == form.cleaned_data[
                    'current_password']:
                    if form.cleaned_data['new_password'] == form.cleaned_data['reenter_new_password']:
                        password = form.cleaned_data['reenter_new_password']
                        if len(password) >= 8:
                            '''
                            passwd = bytes(password, 'utf-8')
                            salt = bcrypt.gensalt(random.randint(8, 12))
                            hashed = bcrypt.hashpw(passwd, salt)
                            '''
                            mentee.password = make_password(password)
                            mentee.save()
                            request.session.flush()
                            return redirect('triple_M:login')
                        else:
                            context = {'form': form, 'reg_no': reg_no,
                                       'error': "Password is too short. It should contain atleast 8 characters"}
                            return render(request, 'triple_M/change_password.html', context)

                    else:
                        context = {'form': form, 'reg_no': reg_no,
                                   'error': "New password entered doesn't match with the reentered password"}
                        return render(request, 'triple_M/change_password.html', context)
                else:
                    context = {'form': form, 'reg_no': reg_no, 'error': "Current Password does not match"}
                    return render(request, 'triple_M/change_password.html', context)
            else:
                form = ChangePassword()
                context = {'form': form, 'reg_no': reg_no}
                return render(request, 'triple_M/change_password.html', context)
        if 'email' in request.session:
            print("YES")
            email = request.session['email']
            form = ChangePassword(request.POST)
            if form.is_valid():
                mentor = models.Mentor.objects.get(mentor_email=email)
                if check_password(form.cleaned_data['current_password'],
                                  mentor.mentor_password) or mentor.mentor_password == form.cleaned_data[
                    'current_password']:
                    if form.cleaned_data['new_password'] == form.cleaned_data['reenter_new_password']:
                        password = form.cleaned_data['reenter_new_password']
                        if len(password) >= 8:
                            '''
                            passwd = bytes(password, 'utf-8')
                            salt = bcrypt.gensalt(random.randint(8, 12))
                            hashed = bcrypt.hashpw(passwd, salt)
                            '''
                            mentor.mentor_password = make_password(password)
                            mentor.save()
                            request.session.flush()
                            return redirect('triple_M:login')
                        else:
                            context = {'form': form, 'email': email,
                                       'error': "Password is too short. It should contain atleast 8 characters"}
                            return render(request, 'triple_M/change_password.html', context)

                    else:
                        context = {'form': form, 'email': email,
                                   'error': "New password entered doesn't match with the reentered password"}
                        return render(request, 'triple_M/change_password.html', context)
                else:
                    context = {'form': form, 'email': email, 'error': "Current Password does not match"}
                    return render(request, 'triple_M/change_password.html', context)
            else:
                context = {'form': form}
                return render(request, 'triple_M/change_password.html', context)


def logout(request):
    request.session.flush()
    return redirect("triple_M:login")


def get_sem_count(year):
    if year == "FE":
        sem_count = 2 * 6

    elif year == "SE":
        sem_count = 4 * 6

    elif year == "TE":
        sem_count = 6 * 6

    elif year == "BE":
        sem_count = 8 * 6

    return sem_count


def check_exam_form(data, count):
    attribute = list(data.keys())
    error = []

    for i in range(0, count):
        if data[attribute[i]] is None:
            error.append(i + 1)

    return error


def check_personal_detail_form(data):
    count = 0
    error = {'profile_photo': [], 'email': [], 'gender': [], 'dob': [], 'ph_no': [], 'parent_ph_no': [], 'address': [],
             'department': [], 'year': [], 'division': [], 'roll_no': []}
    for key in data.keys():
        if data[key] == None:
            error[key].append("Cannot be Empty")
            count += 1

    if count == 0:
        if len(str(data['ph_no'])) is not 10 or str(data['ph_no'])[0] != "7" and str(data['ph_no'])[0] != "8" and \
                str(data['ph_no'])[0] != "9":
            error['ph_no'].append("Invalid Phone Number")
            count += 1

        if len(str(data['parent_ph_no'])) is not 10 or str(data['parent_ph_no'])[0] != "7" and \
                str(data['parent_ph_no'])[0] != "8" and \
                str(data['parent_ph_no'])[0] != "9":
            error['parent_ph_no'].append("Invalid Phone Number")
            count += 1
    if count == 0:
        return 'pass'
    else:
        return error


def validate_mentee_login(regno, password):
    try:
        student = models.Admitted_Student.objects.get(reg_no=regno)
        field_object = models.Admitted_Student._meta.get_field('password')
        legit_password = field_object.value_from_object(student)
        if check_password(password, legit_password) or password == legit_password:
            return "success"
        else:
            return "Password does not match!"
    except:
        error = "Register Number Does not exist!"
        return error


def validate_mentor_login(email, password):
    try:
        mentor = models.Mentor.objects.get(mentor_email=email)
        if check_password(password, mentor.mentor_password) or mentor.mentor_password == password:
            return "success"
        else:
            return "Password does not match!"
    except:
        error = "Email does not exist!"
        return error


def get_avg(dict):
    keys = dict.keys()
    count = 0
    total = 0
    for i in keys:
        if 'sgpi' in i:
            if dict[i] != 0 and dict[i] is not None:
                total += dict[i]
                count += 1
    if count == 0:
        return 0
    else:
        return round(total / count, 2)


def get_kt(dict):
    keys = dict.keys()
    total = 0
    for i in keys:
        if 'atkt' in i:
            if dict[i] != 0 and dict[i] is not None:
                total += dict[i]
    return total


def get_mentee(regno):
    mentee = models.Admitted_Student.objects.get(reg_no=regno).__dict__
    return mentee


def get_mentee_personal_detail(id):
    mentee_personal_detail, created = models.Personal_Detail.objects.get_or_create(
        admitted_student_id=id)
    mentee_personal_detail = mentee_personal_detail.__dict__
    return mentee_personal_detail


def get_mentee_exam_record(id):
    mentee_exam_record, created = models.Exam_Record.objects.get_or_create(admitted_student_id=id)
    mentee_exam_record = mentee_exam_record.__dict__
    del mentee_exam_record['_state']
    del mentee_exam_record['year']
    del mentee_exam_record['id']
    del mentee_exam_record['admitted_student_id']
    return mentee_exam_record


def get_mentor_name(id):
    mentor_name = models.Mentor_Mentee.objects.values('mentor__mentor_name').get(
        admitted_student_id=id)
    mentor_name = mentor_name['mentor__mentor_name']
    return mentor_name


def get_internship_detail(id):
    internship_list = list(models.Internship_Detail.objects.filter(admitted_student_id=id))
    intern_list = []
    for i in internship_list:
        i = i.__dict__
        intern_list.append(i)
    return intern_list


def get_placement_detail(id):
    placement_list = list(models.Placement_Detail.objects.filter(admitted_student_id=id))
    placed_list = []
    for i in placement_list:
        i = i.__dict__
        placed_list.append(i)

    return placed_list


def get_verified_details(list):
    count = 0
    for i in list:
        if i['verified'] == "YES":
            count += 1
    return count


def get_all_verification(mentee_list):
    list_student = []
    count_not_verified = 0
    count_verified = 0
    not_filled = 0
    for i in mentee_list:
        student = {}
        student['reg_no'] = i['reg_no']
        student['first_name'] = i['first_name']
        student['last_name'] = i['last_name']
        try:
            val = models.Personal_Detail.objects.get(admitted_student__reg_no=i['reg_no'])
            if val.email:
                student['personal_detail_verified'] = val.verified
            else:
                student['personal_detail_verified'] = "NOT FILLED"
        except:
            student['personal_detail_verified'] = "NOT FILLED"
        try:
            val = models.Exam_Record.objects.get(admitted_student__reg_no=i['reg_no'])
            if val.sem_1_cleared:
                student['exam_record_verified'] = val.verified
            else:
                student['exam_record_verified'] = "NOT FILLED"
        except:
            student['exam_record_verified'] = "NOT FILLED"

        student['internship_verified'] = verify_placement_internship(
            list(models.Internship_Detail.objects.values("verified").filter(admitted_student__reg_no=i['reg_no'])))

        student['placement_verified'] = verify_placement_internship(
            list(models.Placement_Detail.objects.values("verified").filter(admitted_student__reg_no=i['reg_no'])))

        if student['personal_detail_verified'] == "YES" and student['exam_record_verified'] == "YES":
            count_verified += 1
        elif student['personal_detail_verified'] == "NOT FILLED" or student['exam_record_verified'] == "NOT FILLED":
            not_filled += 1
        else:
            count_not_verified += 1

        list_student.append(student)

    list_student.append(not_filled)
    list_student.append(count_verified)
    list_student.append(count_not_verified)

    return list_student


def verify_placement_internship(list):
    if list:
        for i in list:
            if i['verified'] == 'NO':
                return 'NO'
        return 'YES'
    else:
        return 'NOT FILLED'


def verify_all_details(reg_no):
    student = {}
    try:
        student['personal_detail_verified'] = \
            models.Personal_Detail.objects.values("verified").get(admitted_student__reg_no=reg_no)[
                'verified']
    except:
        student['personal_detail_verified'] = "NO"
    try:
        student['exam_record_verified'] = \
            models.Exam_Record.objects.values("verified").get(admitted_student__reg_no=reg_no)['verified']
    except:
        student['exam_record_verified'] = "NO"

    student['internship_verified'] = verify_placement_internship(
        list(models.Internship_Detail.objects.values("verified").filter(admitted_student__reg_no=reg_no)))

    student['placement_verified'] = verify_placement_internship(
        list(models.Placement_Detail.objects.values("verified").filter(admitted_student__reg_no=reg_no)))

    return student


def export_csv(request, section):
    rowss = []
    if section == "internship_details":
        name = 'internship_details'
        columns = ['Mentee Name', 'Name of Mentor', 'Name of Company', 'Date From', 'Date To', 'Certificate',
                   'Internship Evaluation', 'Class-RollNo', 'Impact Analysis']
        title = "Internship Details"

        rows = models.Internship_Detail.objects.filter(
            admitted_student__mentor_mentee__mentor__mentor_email="shilpa.mhatre@sakec.ac.in"
        ).values_list(
            'admitted_student__first_name',
            'admitted_student__last_name',
            'admitted_student__mentor_mentee__mentor__mentor_name',
            'company',
            'duration_start', 'duration_end', 'certificate', 'intern_eval',
            'admitted_student__personal_detail__year',
            'admitted_student__personal_detail__division',
            'admitted_student__personal_detail__roll_no'
        )
    else:
        name = 'placement_details'
        columns = ['Mentee Name', 'Name of Mentor', 'Name of Company', 'CTC', 'Bond', 'Placed Through',
                   'Offer Letter', 'Class-RollNo', 'Impact Analysis']
        title = "Placement Details"

        rows = models.Placement_Detail.objects.filter(
            admitted_student__mentor_mentee__mentor__mentor_email="shilpa.mhatre@sakec.ac.in"
        ).values_list(
            'admitted_student__first_name',
            'admitted_student__last_name',
            'admitted_student__mentor_mentee__mentor__mentor_name',
            'company',
            'ctc', 'bond', 'placed_through', 'offer_letter',
            'admitted_student__personal_detail__year',
            'admitted_student__personal_detail__division',
            'admitted_student__personal_detail__roll_no'
        )

    for row in rows:
        iteration = ['', '', '', '', '', '', '', '', '']
        iteration[0] = "{0} {1}".format(row[0], row[1])
        iteration[1] = row[2]
        iteration[2] = row[3]
        iteration[3] = row[4]
        iteration[4] = row[5]
        iteration[5] = row[6]
        iteration[6] = row[7]
        iteration[7] = "{0}{1}-{2}".format(row[8], row[9], row[10])
        iteration[8] = "-"
        rowss.append(iteration)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-{name}.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'), name=name
    )

    wb = Workbook()
    worksheet = wb.active
    worksheet.title = title

    row_num = 1

    columns = columns

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)

    for row in rowss:
        row_num += 1
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    wb.save(response)
    return response


def handle_uploaded_file(f):
    f = "intership_file.pdf"
    return f
